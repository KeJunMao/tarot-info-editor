import pandas as pd
import json
import click
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# 水晶名称映射表（JSON英文下划线风格 -> Excel中文）
CRYSTAL_MAP = {
    'clear_quartz': '白水晶',
    'laser_quartz': '激光柱',
    'citrine': '黄水晶',
    'amethyst': '紫水晶',
    'deep_amethyst': '深紫晶',
    'lavender': '薰衣草紫晶',
    'rose_quartz': '粉晶',
    'morganite': '摩根石',
    'strawberry': '草莓晶',
    'rhodochrosite': '红纹石',
    'carnelian': '红玛瑙',
    'garnet': '石榴石',
    'obsidian': '黑曜石',
    'smoky_quartz': '烟晶',
    'hematite': '赤铁矿',
    'tigers_eye': '虎眼石',
    'golden_tiger': '金虎眼',
    'moonstone': '月光石',
    'grey_moon': '灰月光',
    'labradorite': '拉长石',
    'fluorite': '萤石',
    'sunstone': '日光石',
    'amber': '琥珀',
    'rutilated': '金发晶'
}

COLOR_MAP = {
    'pale_yellow': '淡黄',
    'neon_yellow': '霓虹黄',
    'yellow': '黄色',
    'orange': '橙色',
    'blue': '蓝色',
    'silver_white': '银白',
    'green': '绿色',
    'emerald': '祖母绿',
    'red': '红色',
    'crimson': '绯红',
    'red_orange': '红橙',
    'deep_purple': '深紫',
    'light_purple': '浅紫',
    'amber_color': '琥珀色',
    'silver_gray': '银灰',
    'gold': '金色',
    'gray': '灰色',
    'purple': '紫色',
    'royal_blue': '宝蓝',
    'indigo': '靛蓝',
    'navy_blue': '深蓝',
    'sea_green': '海绿',
    'teal': '蓝绿',
    'black': '黑色',
    'iridescent': '虹色',
    'indigo_blue': '靛青',
    'scarlet': '猩红',
    'electric_blue': '电光蓝',
    'silver_red': '银红',
    'golden_yellow': '金黄',
    'gray_white': '灰白',
    'bright_red': '亮红',
    'white': '白色',
    'rust_red': '铁锈红',
    'tender_green': '嫩绿',
    'dark_gray': '深灰',
    'orange_red': '橙红',
    'bright_yellow': '亮黄',
    'silver_purple': '银紫',
    'tan': '黄褐',
    'gold_red': '金红',
    'pink': '粉红',
    'orange_yellow': '橙黄',
    'dark_red': '暗红',
    'copper_green': '铜绿',
    'purple_blue': '紫蓝',
    'dark_brown': '深褐',
    'brown': '褐色',
    'silver_blue': '银蓝',
    'orange_gold': '橙金',
    'light_blue': '淡蓝',
    'bright_blue': '亮蓝',
    'violet': '紫罗兰',
    'iron_red': '铁红',
    'olive_green': '橄榄绿',
    'olive': '橄榄'
}


# 反向映射（Excel中文 -> JSON英文下划线风格）
CRYSTAL_MAP_REVERSE = {v: k for k, v in CRYSTAL_MAP.items()}
COLOR_MAP_REVERSE = {v: k for k, v in COLOR_MAP.items()}

def normalize_crystal_name(name):
    """将各种格式的英文水晶名转为下划线小写格式"""
    if not isinstance(name, str):
        return str(name).lower().replace(' ', '_').replace("'", "")
    return name.lower().replace(' ', '_').replace("'", "")


def json_to_excel(json_file, excel_file):
    # 读取JSON文件
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 确保data是列表
    if not isinstance(data, list):
        data = [data]

    # 提取主表数据
    main_data_list = []
    for card in data:
        main_data_list.append({
            "牌名称": card["label"],
            "牌组": card["suit"],
            "故事": card.get("story", "placeholder story"),
            "图片路径": card["image"],
            "3D图片路径": card["image3d"]
        })
    df_main = pd.DataFrame(main_data_list)

    # 提取元素坐标数据
    element_coords = []
    for card in data:
        for element in card["elements"]:
            element_coords.append({
                "牌名称": card["label"],
                "元素名称": element["label"],
                "x": element.get("x"),
                "y": element.get("y"),
                "r": element.get("r")
            })
    df_element_coords = pd.DataFrame(element_coords)

    # 提取元素详情数据
    elements = []
    for card in data:
        for element in card["elements"]:
            for detail in element["details"]:
                elements.append({
                    "牌名称": card["label"],
                    "元素名称": element["label"],
                    "类型": detail["type"],
                    "内容": detail["content"]
                })
    df_elements = pd.DataFrame(elements)

    # 提取正位逆位含义数据
    meanings = []
    for card in data:
        upright = card["meanings"]["upright"]
        reversed_meaning = card["meanings"]["reversed"]
        meanings.append({
            "牌名称": card["label"],
            "类型": "upright",
            "关键词": ", ".join(keyword.title() for keyword in upright["keywords"]),
            "总结": upright["summary"],
        })
        meanings.append({
            "牌名称": card["label"],
            "类型": "reversed",
            "关键词": ", ".join(keyword.title() for keyword in reversed_meaning["keywords"]),
            "总结": reversed_meaning["summary"],
        })
    df_meanings = pd.DataFrame(meanings)

    # 提取场景数据
    scenarios = []
    for card in data:
        upright = card["meanings"]["upright"]["scenarios"]
        reversed_meaning = card["meanings"]["reversed"]["scenarios"]
        for scenario in upright:
            scenarios.append({
                "牌名称": card["label"],
                "类型": "upright",
                "场景类型": scenario["type"],
                "场景内容": scenario["content"]
            })
        for scenario in reversed_meaning:
            scenarios.append({
                "牌名称": card["label"],
                "类型": "reversed",
                "场景类型": scenario["type"],
                "场景内容": scenario["content"]
            })
    df_scenarios = pd.DataFrame(scenarios)

    # 提取幸运属性数据（带水晶中文转换）
    lucky_data = []
    for card in data:
        lucky = card.get("lucky", {})
        card_label = card["label"]

        # 处理幸运颜色（保持原样）
        for color in lucky.get("colors", ["", ""]):
            lucky_data.append({
                "牌名称": card_label,
                "属性类型": "颜色",
                "属性值": color
            })

        # 处理幸运数字（保持原样）
        for number in lucky.get("numbers", ["", ""]):
            lucky_data.append({
                "牌名称": card_label,
                "属性类型": "数字",
                "属性值": number
            })

        # 处理幸运水晶（英文转为中文）
        for crystal in lucky.get("crystals", ["", ""]):
            crystal_key = normalize_crystal_name(crystal)
            crystal_cn = CRYSTAL_MAP.get(crystal_key, crystal)  # 找不到映射保持原值
            lucky_data.append({
                "牌名称": card_label,
                "属性类型": "水晶",
                "属性值": crystal_cn
            })

    df_lucky = pd.DataFrame(lucky_data)

    # 写入Excel文件
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df_main.to_excel(writer, sheet_name="主表", index=False)
        df_element_coords.to_excel(writer, sheet_name="元素", index=False)
        df_elements.to_excel(writer, sheet_name="元素详情", index=False)
        df_meanings.to_excel(writer, sheet_name="正位逆位含义", index=False)
        df_scenarios.to_excel(writer, sheet_name="场景表", index=False)
        df_lucky.to_excel(writer, sheet_name="幸运属性", index=False)

    # 添加数据验证约束
    wb = load_workbook(excel_file)

    # 获取主表行数
    max_row_main = len(df_main) + 1

    # 获取所有唯一的牌组
    suits = df_main["牌组"].unique().tolist()
    suits_str = ','.join(suits)

    # 主表约束
    ws_main = wb["主表"]

    # 牌组：下拉列表
    dv_suit = DataValidation(type="list", formula1=f'"{suits_str}"', allow_blank=False, showErrorMessage=True)
    dv_suit.error = '牌组必须从列表中选择'
    dv_suit.errorTitle = '无效的牌组'
    dv_suit.add(f'B2:B{max_row_main}')
    ws_main.add_data_validation(dv_suit)

    # 元素表约束
    ws_elements = wb["元素"]
    max_row_elem = len(df_element_coords) + 1

    # 牌名称：下拉列表
    dv_elem_card = DataValidation(type="list", formula1='主表!$A$2:$A$' + str(max_row_main), allow_blank=False, showErrorMessage=True)
    dv_elem_card.error = '牌名称必须从主表选择'
    dv_elem_card.errorTitle = '无效的牌名称'
    dv_elem_card.add(f'A2:A{max_row_elem}')
    ws_elements.add_data_validation(dv_elem_card)

    # x, y, r：数字且在合理范围（0-1000）
    dv_coord = DataValidation(type="whole", operator="between", formula1='0', formula2='1000', allow_blank=True, showErrorMessage=True)
    dv_coord.error = '坐标值必须是0-1000之间的数字'
    dv_coord.errorTitle = '无效的坐标'
    dv_coord.add(f'C2:E{max_row_elem}')
    ws_elements.add_data_validation(dv_coord)

    # 元素详情表约束
    ws_details = wb["元素详情"]
    max_row_details = len(df_elements) + 1

    # 牌名称：下拉列表
    dv_detail_card = DataValidation(type="list", formula1='主表!$A$2:$A$' + str(max_row_main), allow_blank=False, showErrorMessage=True)
    dv_detail_card.error = '牌名称必须从主表选择'
    dv_detail_card.add(f'A2:A{max_row_details}')
    ws_details.add_data_validation(dv_detail_card)

    # 类型：下拉列表（visual, symbolism, interpretation）
    dv_detail_type = DataValidation(type="list", formula1='"visual,symbolism,interpretation"', allow_blank=False, showErrorMessage=True)
    dv_detail_type.error = '类型必须是：visual、symbolism或interpretation'
    dv_detail_type.errorTitle = '无效的详情类型'
    dv_detail_type.add(f'C2:C{max_row_details}')
    ws_details.add_data_validation(dv_detail_type)

    # 正位逆位含义表约束
    ws_meanings = wb["正位逆位含义"]
    max_row_meanings = len(df_meanings) + 1

    # 牌名称：下拉列表
    dv_mean_card = DataValidation(type="list", formula1='主表!$A$2:$A$' + str(max_row_main), allow_blank=False, showErrorMessage=True)
    dv_mean_card.error = '牌名称必须从主表选择'
    dv_mean_card.add(f'A2:A{max_row_meanings}')
    ws_meanings.add_data_validation(dv_mean_card)

    # 类型：只能是upright或reversed
    dv_mean_type = DataValidation(type="list", formula1='"upright,reversed"', allow_blank=False, showErrorMessage=True)
    dv_mean_type.error = '类型必须是：upright或reversed'
    dv_mean_type.errorTitle = '无效的含义类型'
    dv_mean_type.add(f'B2:B{max_row_meanings}')
    ws_meanings.add_data_validation(dv_mean_type)

    # 场景表约束
    ws_scenarios = wb["场景表"]
    max_row_scenarios = len(df_scenarios) + 1

    # 牌名称：下拉列表
    dv_scen_card = DataValidation(type="list", formula1='主表!$A$2:$A$' + str(max_row_main), allow_blank=False, showErrorMessage=True)
    dv_scen_card.error = '牌名称必须从主表选择'
    dv_scen_card.add(f'A2:A{max_row_scenarios}')
    ws_scenarios.add_data_validation(dv_scen_card)

    # 类型：只能是upright或reversed
    dv_scen_type = DataValidation(type="list", formula1='"upright,reversed"', allow_blank=False, showErrorMessage=True)
    dv_scen_type.error = '类型必须是：upright或reversed'
    dv_scen_type.errorTitle = '无效的场景类型'
    dv_scen_type.add(f'B2:B{max_row_scenarios}')
    ws_scenarios.add_data_validation(dv_scen_type)

    # 幸运属性表约束
    ws_lucky = wb["幸运属性"]
    max_row_lucky = len(df_lucky) + 1
    # 牌名称：下拉列表
    dv_lucky_card = DataValidation(type="list", formula1='主表!$A$2:$A$' + str(max_row_main), allow_blank=False, showErrorMessage=True)
    dv_lucky_card.error = '牌名称必须从主表选择'
    dv_lucky_card.errorTitle = '无效的牌名称'
    dv_lucky_card.add(f'A2:A{max_row_lucky}')
    ws_lucky.add_data_validation(dv_lucky_card)

    # 属性类型：下拉列表（颜色, 数字, 水晶）
    dv_lucky_type = DataValidation(type="list", formula1='"颜色,数字,水晶"', allow_blank=False, showErrorMessage=True)
    dv_lucky_type.error = '属性类型必须是：颜色、数字或水晶'
    dv_lucky_type.errorTitle = '无效的属性类型'
    dv_lucky_type.add(f'B2:B{max_row_lucky}')
    ws_lucky.add_data_validation(dv_lucky_type)

    # 为特定类型添加值的下拉列表（颜色和水晶）
    color_options = ','.join(COLOR_MAP.values())
    crystal_options = ','.join(CRYSTAL_MAP.values())

    for row in range(2, max_row_lucky + 1):
        cell_type = ws_lucky.cell(row=row, column=2).value  # B列是属性类型

        if cell_type == "颜色":
            dv_color = DataValidation(
                type="list",
                formula1=f'"{color_options}"',
                allow_blank=False,
                showErrorMessage=True
            )
            dv_color.error = '颜色必须从预设列表中选择（中文）'
            dv_color.errorTitle = '无效的颜色名称'
            dv_color.add(f'C{row}')  # C列是属性值
            ws_lucky.add_data_validation(dv_color)

        elif cell_type == "水晶":
            dv_crystal = DataValidation(
                type="list",
                formula1=f'"{crystal_options}"',
                allow_blank=False,
                showErrorMessage=True
            )
            dv_crystal.error = '水晶必须从预设列表中选择（中文）'
            dv_crystal.errorTitle = '无效的水晶名称'
            dv_crystal.add(f'C{row}')  # C列是属性值
            ws_lucky.add_data_validation(dv_crystal)


    wb.save(excel_file)
    return True


def excel_to_json(excel_file, json_file):
    # 读取Excel文件的各表
    df_main = pd.read_excel(excel_file, sheet_name="主表")
    df_element_coords = pd.read_excel(excel_file, sheet_name="元素")
    df_elements = pd.read_excel(excel_file, sheet_name="元素详情")
    df_meanings = pd.read_excel(excel_file, sheet_name="正位逆位含义")
    df_scenarios = pd.read_excel(excel_file, sheet_name="场景表")
    df_lucky = pd.read_excel(excel_file, sheet_name="幸运属性")

    # 转换所有卡片
    cards = []
    main_records = df_main.to_dict(orient="records")

    for main_data in main_records:
        card_name = main_data["牌名称"]
        card_name_lower = card_name.lower() if pd.notna(card_name) else ""

        # 转换元素数据（包含坐标）
        elements = []
        processed_elements = set()
        card_element_coords = df_element_coords[df_element_coords["牌名称"].str.lower() == card_name_lower]
        card_elements = df_elements[df_elements["牌名称"].str.lower() == card_name_lower]

        for index, coord_row in card_element_coords.iterrows():
            element_name = coord_row["元素名称"]
            if element_name not in processed_elements:
                element = {
                    "label": element_name,
                    "x": coord_row["x"],
                    "y": coord_row["y"],
                    "r": coord_row["r"],
                    "details": []
                }
                # 查找对应元素的所有详情
                for detail_row in card_elements[card_elements["元素名称"] == element_name].iterrows():
                    element["details"].append({
                        "type": detail_row[1]["类型"],
                        "content": detail_row[1]["内容"]
                    })
                elements.append(element)
                processed_elements.add(element_name)

        # 转换正位逆位含义数据（添加空值检查）
        card_meanings = df_meanings[df_meanings["牌名称"].str.lower().str.strip() == card_name_lower]

        upright_data = card_meanings[card_meanings["类型"] == "upright"]
        reversed_data = card_meanings[card_meanings["类型"] == "reversed"]

        meanings = {
            "upright": {
                "keywords": upright_data["关键词"].values[0].split(", ") if len(upright_data) > 0 and pd.notna(upright_data["关键词"].values[0]) else [],
                "summary": upright_data["总结"].values[0] if len(upright_data) > 0 and pd.notna(upright_data["总结"].values[0]) else "",
                "scenarios": []
            },
            "reversed": {
                "keywords": reversed_data["关键词"].values[0].split(", ") if len(reversed_data) > 0 and pd.notna(reversed_data["关键词"].values[0]) else [],
                "summary": reversed_data["总结"].values[0] if len(reversed_data) > 0 and pd.notna(reversed_data["总结"].values[0]) else "",
                "scenarios": []
            }
        }

        # 转换场景数据
        card_scenarios = df_scenarios[df_scenarios["牌名称"].str.lower().str.strip() == card_name_lower]
        for index, row in card_scenarios.iterrows():
            if row["类型"] == "upright":
                meanings["upright"]["scenarios"].append({
                    "type": row["场景类型"],
                    "content": row["场景内容"]
                })
            else:
                meanings["reversed"]["scenarios"].append({
                    "type": row["场景类型"],
                    "content": row["场景内容"]
                })

        # 转换幸运属性数据（中文转英文下划线）
        card_lucky = df_lucky[df_lucky["牌名称"].str.lower().str.strip() == card_name_lower]

        colors = []
        numbers = []
        crystals = []

        for index, row in card_lucky.iterrows():
            attr_type = row["属性类型"]
            attr_value = row["属性值"]

            if pd.notna(attr_type) and pd.notna(attr_value):
                if attr_type == "颜色":
                    # 中文转英文下划线
                    color_cn = str(attr_value).strip()
                    color_en = COLOR_MAP_REVERSE.get(color_cn, color_cn.lower().replace(' ', '_'))
                elif attr_type == "数字":
                    # 尝试转换为整数，如果失败则保持原样
                    try:
                        numbers.append(int(float(attr_value)))
                    except (ValueError, TypeError):
                        numbers.append(str(attr_value))
                elif attr_type == "水晶":
                    # 中文转英文下划线风格
                    crystal_cn = str(attr_value).strip()
                    crystal_en = CRYSTAL_MAP_REVERSE.get(crystal_cn, crystal_cn.lower().replace(' ', '_'))
                    crystals.append(crystal_en)

        lucky = {
            "colors": colors,
            "numbers": numbers,
            "crystals": crystals
        }

        # 构建当前卡片的JSON数据
        json_card = {
            "label": main_data["牌名称"],
            "suit": main_data["牌组"],
            "story": main_data["故事"],
            "image": main_data["图片路径"],
            "image3d": main_data["3D图片路径"],
            "elements": elements,
            "meanings": meanings,
            "lucky": lucky
        }
        cards.append(json_card)

    # 写入JSON文件
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(cards, f, indent=4, ensure_ascii=False)

    return True
# json2excel命令
@click.command(name='json2excel')
@click.option('--input', type=click.Path(exists=True, readable=True), required=True, help='Input JSON file path')
@click.option('--output', type=click.Path(writable=True), required=True, help='Output Excel file path')
def json2excel_cmd(input, output):
    success = json_to_excel(input, output)
    if success:
        click.echo(f"Successfully converted {input} to {output}.")
    else:
        click.echo(f"Failed to convert {input} to {output}.")


# excel2json命令
@click.command(name='excel2json')
@click.option('--input', type=click.Path(exists=True, readable=True), required=True, help='Input Excel file path')
@click.option('--output', type=click.Path(writable=True), required=True, help='Output JSON file path')
def excel2json_cmd(input, output):
    success = excel_to_json(input, output)
    if success:
        click.echo(f"Successfully converted {input} to {output}.")
    else:
        click.echo(f"Failed to convert {input} to {output}.")


# 定义命令组
@click.group()
def cli():
    pass


# 添加命令到命令组
cli.add_command(json2excel_cmd)
cli.add_command(excel2json_cmd)


# 运行
if __name__ == "__main__":
    cli()
